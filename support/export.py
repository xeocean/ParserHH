import json
import os
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def export_db_to_excel(db_name: str, table_name: str = "vacancies"):
    try:
        db_path = Path(db_name)
        if db_path.exists() and db_path.is_file():

            export_dir = Path("export")
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)

            excel_file_name = input("Введите имя файла: ")

            export_path = os.path.join(export_dir, f"{excel_file_name}.xlsx")
            with sqlite3.connect(db_name) as conn:
                cursor = conn.cursor()

                # Получаем все данные из таблицы
                cursor.execute(f"SELECT * FROM {table_name}")
                rows = cursor.fetchall()

                # Получаем имена столбцов
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = [col[1] for col in cursor.fetchall()]

                # Создаем новый Excel файл
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = table_name

                # Записываем имена столбцов в первую строку
                for col_num, column_title in enumerate(columns, 1):
                    cell = sheet.cell(row=1, column=col_num, value=column_title)

                    # Устанавливаем ширину столбца по умолчанию 20
                    sheet.column_dimensions[get_column_letter(col_num)].width = 20

                    # Устанавливаем ширину 60 для столбцов F, H и I
                    if col_num in [6, 8, 9]:
                        sheet.column_dimensions[get_column_letter(col_num)].width = 60

                    # Включаем перенос текста для заголовков столбцов
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                # Записываем строки данных
                for row_num, row_data in enumerate(rows, 2):
                    for col_num, cell_data in enumerate(row_data, 1):
                        cell = sheet.cell(row=row_num, column=col_num)

                        if columns[col_num - 1] == "url":
                            cell.hyperlink = cell_data
                            cell.value = "Ссылка на вакансию"
                            cell.font = Font(color="0000FF", underline="single")
                        else:
                            cell.value = cell_data

                        # Устанавливаем перенос строк для текстовых ячеек
                        if isinstance(cell_data, str):
                            cell.value = cell_data
                            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                # Сохраняем Excel файл
                workbook.save(export_path)
                print(f"Данные экспортированы в {export_path}")

        else:
            print("Файл базы данных не найден или не является файлом")

    except Exception as e:
        print(f"Ошибка: {e}")
        return


def export_db_to_json(db_name: str, table_name: str = "vacancies"):
    try:
        db_path = Path(db_name)
        if db_path.exists() and db_path.is_file():

            export_dir = Path("export")
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)

            json_file_name = input("Введите имя файла: ")

            export_path = os.path.join(export_dir, f"{json_file_name}.json")
            with sqlite3.connect(db_name) as conn:
                cursor = conn.cursor()

                # Получаем все данные из таблицы
                cursor.execute(f"SELECT * FROM {table_name}")
                values = cursor.fetchall()

                # Получаем имена столбцов
                cursor.execute(f"PRAGMA table_info({table_name})")
                keys = [col[1] for col in cursor.fetchall()]

                # Преобразуем данные в список словарей
                result = [dict(zip(keys, row)) for row in values]

                # Записываем результат в JSON-файл
                with open(export_path, "w", encoding="utf-8") as f:
                    json.dump(result, f, ensure_ascii=False, indent=4)

                print(f"Данные успешно экспортированы в {export_path}")

        else:
            print("Файл базы данных не найден или не является файлом")

    except Exception as e:
        print(f"Ошибка: {e}")
        return
